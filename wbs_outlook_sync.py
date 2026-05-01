#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WBS <-> Outlook カレンダー同期ツール

初回セットアップ（順番通りに実行）:
  1. python wbs_outlook_sync.py add-time-cols  --file WBS.xlsx
  2. python wbs_outlook_sync.py add-status-col --file WBS.xlsx

通常運用:
  python wbs_outlook_sync.py wbs2outlook  --file WBS.xlsx --id user@example.com
  python wbs_outlook_sync.py outlook2wbs --file WBS.xlsx --id user@example.com

新規WBS作成:
  python wbs_outlook_sync.py new-wbs --output new.xlsx --gantt-start 2026-05-01 --gantt-end 2026-07-31
"""

import argparse
import re
import sys
from datetime import date, datetime, time, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() != "utf-8":
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

try:
    import win32com.client
    OUTLOOK_AVAILABLE = True
except ImportError:
    OUTLOOK_AVAILABLE = False


# ─── 列定義（add-time-cols + add-status-col 適用後）──────────────────────
COL_PROJECT_NAME  = 1
COL_PROJECT_ID    = 2
COL_PHASE         = 3
COL_L1            = 4
COL_L2            = 5
COL_L3            = 6
COL_L4_TASK       = 7
COL_TASK_ID       = 8
COL_ASSIGNED      = 9
COL_AP_MAIL       = 10
COL_START_DATE_P  = 11   # K  開始予定日
COL_START_TIME_P  = 12   # L  開始予定時刻
COL_FINISH_DATE_P = 13   # M  終了予定日
COL_FINISH_TIME_P = 14   # N  終了予定時刻
COL_START_DATE    = 15   # O  開始実績 ← Outlookから書き戻す
COL_FINISH_DATE   = 16   # P  終了実績 ← Outlookから書き戻す
COL_STATUS        = 17   # Q  ステータス（自動計算式）
COL_GANTT_FIRST   = 18   # R〜 ガントチャート

FIXED_HEADERS = [
    "ProjectName", "ProjectID", "Phese", "L1_Category",
    "L2_Major_Classification", "L3_Middle_Classification", "L4_Task",
    "TaskID", "Assinged Person", "AP_MailAddress",
    "StartDate_p", "StartTime_p", "FinishDate_p", "FinishTime_p",
    "StartDate", "FinishDate", "Status",
]

STATUS_DONE        = "完了"
STATUS_IN_PROGRESS = "進行中"
STATUS_NOT_STARTED = "未着手"

WBS_TAG          = "[WBS]"
CANCELLED_PREFIX = "[削除済]"


# ─── 数式生成 ────────────────────────────────────────────────────────────

def _status_formula(row: int) -> str:
    """StartDate/FinishDate から Status を自動判定する Excel 数式"""
    o = get_column_letter(COL_START_DATE)
    p = get_column_letter(COL_FINISH_DATE)
    return (
        f'=IF(AND({o}{row}<>"",{p}{row}<>""),"{STATUS_DONE}",'
        f'IF({o}{row}<>"","{STATUS_IN_PROGRESS}","{STATUS_NOT_STARTED}"))'
    )


def _gantt_formula(col: int, row: int) -> str:
    """ガントチャートのセル数式"""
    c = get_column_letter(col)
    k = get_column_letter(COL_START_DATE_P)
    m = get_column_letter(COL_FINISH_DATE_P)
    return f'=IF(AND({c}$1>=${k}{row},{c}$1<=${m}{row}),1,"")'


def _regen_gantt(ws, gantt_first: int) -> int:
    """gantt_first 以降のガント数式を現在の列位置で再生成。更新数を返す。"""
    count = 0
    for row in range(2, ws.max_row + 1):
        for col in range(gantt_first, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value is not None:
                cell.value = _gantt_formula(col, row)
                count += 1
    return count


# ─── Outlook 件名キー（ProjectID + TaskID の複合キー）──────────────────

def make_subject(project_id: str, task_id, task_name: str) -> str:
    """例: [WBS][D-0001][1] システム要件記入"""
    return f"{WBS_TAG}[{project_id}][{task_id}] {task_name}"


def parse_key(subject: str):
    """
    件名から (project_id: str, task_id: int) を返す。
    [削除済] プレフィックスのもの・WBSタグのないものは None。
    """
    s = subject or ""
    if s.startswith(CANCELLED_PREFIX):
        return None
    m = re.search(r'\[WBS\]\[([^\]]+)\]\[(\d+)\]', s)
    return (m.group(1), int(m.group(2))) if m else None


# ─── ユーティリティ ──────────────────────────────────────────────────────

def combine_dt(d, t) -> datetime | None:
    """date + time → datetime。時刻未設定は 09:00 とする。"""
    if d is None:
        return None
    d = d.date() if isinstance(d, datetime) else d
    if isinstance(t, datetime):
        t = t.time()
    return datetime.combine(d, t if isinstance(t, time) else time(9, 0))


def pytime_to_dt(pt) -> datetime:
    """pywintypes.datetime → Python datetime"""
    return datetime(pt.year, pt.month, pt.day, pt.hour, pt.minute, pt.second)


def is_completed(task: dict) -> bool:
    """StartDate と FinishDate の両方が入っていれば完了"""
    return task["start_date"] is not None and task["finish_date"] is not None


def _require_outlook():
    if not OUTLOOK_AVAILABLE:
        print("エラー: pywin32 が必要です。pip install pywin32 を実行してください。",
              file=sys.stderr)
        sys.exit(1)


def _open_calendar():
    _require_outlook()
    ol = win32com.client.Dispatch("Outlook.Application")
    ns = ol.GetNamespace("MAPI")
    return ol, ns.GetDefaultFolder(9)  # 9 = olFolderCalendar


def _load_tasks(ws, target_id: str) -> list[dict]:
    """AP_MailAddress が target_id に一致するタスク行を辞書リストで返す"""
    result = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        ap = row[COL_AP_MAIL - 1]
        if not ap or str(ap).strip().lower() != target_id.strip().lower():
            continue
        result.append({
            "row":           row_idx,
            "project_name":  row[COL_PROJECT_NAME - 1],
            "project_id":    str(row[COL_PROJECT_ID - 1]) if row[COL_PROJECT_ID - 1] else None,
            "phase":         row[COL_PHASE - 1],
            "l4_task":       row[COL_L4_TASK - 1],
            "task_id":       row[COL_TASK_ID - 1],
            "assigned":      row[COL_ASSIGNED - 1],
            "ap_mail":       ap,
            "start_date_p":  row[COL_START_DATE_P - 1],
            "start_time_p":  row[COL_START_TIME_P - 1],
            "finish_date_p": row[COL_FINISH_DATE_P - 1],
            "finish_time_p": row[COL_FINISH_TIME_P - 1],
            "start_date":    row[COL_START_DATE - 1],
            "finish_date":   row[COL_FINISH_DATE - 1],
        })
    return result


def _scan_events(calendar) -> dict[tuple, object]:
    """
    カレンダーから [WBS][ProjectID][TaskID] のアクティブイベントを収集。
    [削除済] のものは除外。複合キー (project_id, task_id) でインデックス。
    """
    events: dict[tuple, object] = {}
    for item in calendar.Items:
        try:
            key = parse_key(item.Subject)
            if key is not None:
                events[key] = item
        except Exception:
            pass
    return events


def _mark_cancelled(apt) -> None:
    """Outlookイベントを論理削除: 件名に [削除済] プレフィックスを付与"""
    if not apt.Subject.startswith(CANCELLED_PREFIX):
        apt.Subject = CANCELLED_PREFIX + apt.Subject
        apt.Save()


# ─── wbs2outlook ────────────────────────────────────────────────────────

def cmd_wbs2outlook(args):
    """
    WBS の予定日程を Outlook カレンダーに登録/更新する。

    - 完了済みタスク（StartDate + FinishDate 両方あり）はスキップし、
      既存Outlookイベントがあれば論理削除する。
    - WBSから消えたタスクの Outlook イベントも論理削除する。
    - マッチングキー: ProjectID + TaskID の複合キー
    """
    wb = openpyxl.load_workbook(args.file, data_only=True)
    tasks = _load_tasks(wb.active, args.id)

    if not tasks:
        print(f"対象タスクなし (AP_MailAddress: {args.id})")
        return

    print(f"対象タスク: {len(tasks)}件  (ID: {args.id})")
    ol, cal = _open_calendar()
    existing = _scan_events(cal)
    active: set[tuple] = set()
    created = updated = cancelled = skipped = 0

    for t in tasks:
        pid, tid = t["project_id"], t["task_id"]
        label = f"[{pid}][{tid}] {t['l4_task']}"

        if pid is None or tid is None:
            print(f"  スキップ (ProjectID/TaskID未設定): 行{t['row']}")
            skipped += 1
            continue

        # 完了済みは active に入れない → 論理削除対象になる
        if is_completed(t):
            print(f"  スキップ (完了): {label}")
            skipped += 1
            continue

        start = combine_dt(t["start_date_p"], t["start_time_p"])
        end   = combine_dt(t["finish_date_p"], t["finish_time_p"])

        if start is None or end is None:
            print(f"  スキップ (予定日未設定): {label}")
            skipped += 1
            continue

        if end <= start:
            print(f"  スキップ (終了日≤開始日): {label}")
            skipped += 1
            continue

        key  = (pid, int(tid))
        active.add(key)
        subj = make_subject(pid, tid, t["l4_task"] or "")
        body = (
            f"TaskID: {tid}\n"
            f"Project: {t['project_name']}  ({pid})\n"
            f"Phase: {t['phase']}\n"
            f"Task: {t['l4_task']}\n"
            f"Assigned: {t['assigned']}"
        )
        fmt = "%Y/%m/%d %H:%M"

        if key in existing:
            apt = existing[key]
            apt.Subject = subj
            apt.Start   = start.strftime(fmt)
            apt.End     = end.strftime(fmt)
            apt.Body    = body
            apt.Save()
            print(f"  更新: {label}  {start:%m/%d %H:%M}–{end:%m/%d %H:%M}")
            updated += 1
        else:
            apt             = ol.CreateItem(1)  # olAppointmentItem
            apt.Subject     = subj
            apt.Start       = start.strftime(fmt)
            apt.End         = end.strftime(fmt)
            apt.Body        = body
            apt.AllDayEvent = False
            apt.Save()
            print(f"  作成: {label}  {start:%m/%d %H:%M}–{end:%m/%d %H:%M}")
            created += 1

    # WBSにない（消えた/完了した）イベントを論理削除
    for key, apt in existing.items():
        if key not in active:
            _mark_cancelled(apt)
            print(f"  論理削除: [{key[0]}][{key[1]}]  → {CANCELLED_PREFIX} を付与")
            cancelled += 1

    print(f"\n完了: 作成={created}  更新={updated}  論理削除={cancelled}  スキップ={skipped}")


# ─── outlook2wbs ────────────────────────────────────────────────────────

def cmd_outlook2wbs(args):
    """
    Outlook カレンダーのイベント日程を WBS の実績欄（StartDate/FinishDate）に書き戻す。

    - 完了済みタスクはスキップ（上書き禁止）。
    - [削除済] イベントはスキップ。
    - マッチングキー: ProjectID + TaskID の複合キー
    """
    wb = openpyxl.load_workbook(args.file)
    ws = wb.active
    tasks = _load_tasks(ws, args.id)

    if not tasks:
        print(f"対象タスクなし (AP_MailAddress: {args.id})")
        return

    print(f"対象タスク: {len(tasks)}件  (ID: {args.id})")
    _, cal = _open_calendar()

    # 完了済みを除いた対象の複合キーセット
    target_keys = {
        (t["project_id"], int(t["task_id"]))
        for t in tasks
        if t["task_id"] is not None
        and t["project_id"] is not None
        and not is_completed(t)
    }

    # Outlook から対象イベントのみ取得
    ol_ev: dict[tuple, dict] = {}
    for item in cal.Items:
        try:
            key = parse_key(item.Subject)
            if key in target_keys:
                ol_ev[key] = {
                    "start": pytime_to_dt(item.Start),
                    "end":   pytime_to_dt(item.End),
                }
        except Exception:
            pass

    updated = skipped = 0
    for t in tasks:
        pid, tid = t["project_id"], t["task_id"]
        label = f"[{pid}][{tid}] {t['l4_task']}"

        if pid is None or tid is None:
            skipped += 1
            continue

        if is_completed(t):
            print(f"  スキップ (完了済・上書き禁止): {label}")
            skipped += 1
            continue

        key = (pid, int(tid))
        if key not in ol_ev:
            print(f"  スキップ (Outlookに未登録): {label}")
            skipped += 1
            continue

        ev = ol_ev[key]
        ws.cell(t["row"], COL_START_DATE).value  = ev["start"]
        ws.cell(t["row"], COL_FINISH_DATE).value = ev["end"]
        print(f"  更新: {label}  {ev['start']:%Y-%m-%d %H:%M}–{ev['end']:%Y-%m-%d %H:%M}")
        updated += 1

    out = Path(args.output) if args.output else Path(args.file)
    wb.save(out)
    print(f"\n完了: 更新={updated}  スキップ={skipped}  →  {out}")


# ─── add-time-cols ───────────────────────────────────────────────────────

def cmd_add_time_cols(args):
    """
    既存WBSに StartTime_p (列L) と FinishTime_p (列N) を挿入する。
    ガントチャートの数式を再生成する。初回のみ実行。
    """
    wb = openpyxl.load_workbook(args.file)
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "StartTime_p" in headers:
        print("StartTime_p 列はすでに存在します。スキップします。")
        return

    ws.insert_cols(12)
    ws.cell(1, 12).value = "StartTime_p"

    ws.insert_cols(14)
    ws.cell(1, 14).value = "FinishTime_p"

    # この時点のガント開始は col 17（Status未追加）
    n = _regen_gantt(ws, gantt_first=17)

    out = Path(args.output) if args.output else Path(args.file)
    wb.save(out)
    print(f"時間列を追加しました → {out}")
    print("  列L: StartTime_p  /  列N: FinishTime_p  (入力形式: HH:MM)")
    print(f"  ガント数式 {n}件 再生成")
    print(f"\n次のステップ: python wbs_outlook_sync.py add-status-col --file {out}")


# ─── add-status-col ──────────────────────────────────────────────────────

def cmd_add_status_col(args):
    """
    既存WBSに Status 列 (列Q) を FinishDate (列P) の直後に挿入する。
    Status は StartDate/FinishDate から自動計算する Excel 数式。
    ガントチャートは列R (18) 以降にシフトし、数式を再生成する。
    初回のみ実行。
    """
    wb = openpyxl.load_workbook(args.file)
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "Status" in headers:
        print("Status 列はすでに存在します。スキップします。")
        return
    if "StartTime_p" not in headers:
        print("エラー: 先に add-time-cols を実行してください（ステップ1）。")
        sys.exit(1)

    # FinishDate (P=16) の直後に Status を挿入
    ws.insert_cols(17)
    ws.cell(1, 17).value = "Status"

    # TaskID のある行にのみ Status 数式を設定
    formula_count = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, COL_TASK_ID).value is not None:
            ws.cell(row, COL_STATUS).value = _status_formula(row)
            formula_count += 1

    # ガントは col 18 以降になったので再生成
    n = _regen_gantt(ws, gantt_first=COL_GANTT_FIRST)

    out = Path(args.output) if args.output else Path(args.file)
    wb.save(out)
    print(f"Status列を追加しました → {out}")
    print(f"  列Q(17): Status  自動計算数式 {formula_count}行に設定")
    print(f"  ガント列: R(18)〜  数式 {n}件 再生成")
    print()
    print("  Status の自動判定ルール:")
    print(f"    StartDate あり + FinishDate あり  →  {STATUS_DONE}")
    print(f"    StartDate あり + FinishDate なし  →  {STATUS_IN_PROGRESS}")
    print(f"    StartDate なし                    →  {STATUS_NOT_STARTED}")


# ─── new-wbs ─────────────────────────────────────────────────────────────

def cmd_new_wbs(args):
    """
    ガントチャート範囲を指定して新規WBSテンプレートを作成する。
    列構成は add-time-cols + add-status-col 適用済みの状態で生成する。
    """
    try:
        gantt_start = datetime.strptime(args.gantt_start, "%Y-%m-%d").date()
        gantt_end   = datetime.strptime(args.gantt_end,   "%Y-%m-%d").date()
    except ValueError:
        print("エラー: 日付は YYYY-MM-DD 形式で指定してください。")
        sys.exit(1)

    if gantt_end < gantt_start:
        print("エラー: gantt-end は gantt-start より後の日付を指定してください。")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for i, h in enumerate(FIXED_HEADERS, start=1):
        ws.cell(1, i).value = h

    d, col = gantt_start, COL_GANTT_FIRST
    while d <= gantt_end:
        ws.cell(1, col).value = datetime(d.year, d.month, d.day)
        d   += timedelta(days=1)
        col += 1

    wb.save(args.output)
    days = (gantt_end - gantt_start).days + 1
    last = get_column_letter(COL_GANTT_FIRST + days - 1)
    print(f"新規WBSを作成しました → {args.output}")
    print(f"  固定列: A〜Q ({len(FIXED_HEADERS)}列)")
    print(f"  ガント: R〜{last}  ({gantt_start} 〜 {gantt_end}  {days}日間)")
    print()
    print("  データ行追加時に設定が必要な数式:")
    print(f'    Status列(Q): {_status_formula(2)}')
    print(f'    ガント列(R〜): {_gantt_formula(COL_GANTT_FIRST, 2)}')


# ─── メイン ──────────────────────────────────────────────────────────────

def _build_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(
        prog="wbs_outlook_sync",
        description="WBS <-> Outlook カレンダー同期ツール",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
【初回セットアップ（順番通りに実行）】
  1. python wbs_outlook_sync.py add-time-cols  --file WBS.xlsx
  2. python wbs_outlook_sync.py add-status-col --file WBS.xlsx

【通常運用】
  python wbs_outlook_sync.py wbs2outlook  --file WBS.xlsx --id user@example.com
  python wbs_outlook_sync.py outlook2wbs --file WBS.xlsx --id user@example.com [--output WBS_updated.xlsx]

【新規作成】
  python wbs_outlook_sync.py new-wbs --output new.xlsx --gantt-start 2026-05-01 --gantt-end 2026-07-31

【同期ルール】
  - マッチングキー   : ProjectID + TaskID の複合キー
  - Outlook件名形式  : [WBS][ProjectID][TaskID] タスク名
  - 完了タスク       : StartDate + FinishDate 両方あり → 同期スキップ
  - 論理削除         : WBSから消えた/完了したタスクの件名に [削除済] を付与
""",
    )
    sub = ap.add_subparsers(dest="command", required=True)

    for name, help_text in [
        ("add-time-cols",  "既存WBSに時間列を追加【初回のみ / ステップ1】"),
        ("add-status-col", "既存WBSにStatus列を追加【初回のみ / ステップ2】"),
    ]:
        p = sub.add_parser(name, help=help_text)
        p.add_argument("--file",   required=True,  help="対象WBSファイル (.xlsx)")
        p.add_argument("--output", default=None,   help="出力先（省略時は上書き）")

    p = sub.add_parser("wbs2outlook", help="WBS予定日程 → Outlookイベント 登録/更新")
    p.add_argument("--file", required=True, help="対象WBSファイル (.xlsx)")
    p.add_argument("--id",   required=True, metavar="AP_MAIL",
                   help="処理対象の AP_MailAddress")

    p = sub.add_parser("outlook2wbs", help="Outlookイベント → WBS実績 反映")
    p.add_argument("--file",   required=True, help="対象WBSファイル (.xlsx)")
    p.add_argument("--id",     required=True, metavar="AP_MAIL",
                   help="処理対象の AP_MailAddress")
    p.add_argument("--output", default=None, help="出力先（省略時は上書き）")

    p = sub.add_parser("new-wbs", help="新規WBSテンプレート作成")
    p.add_argument("--output",      required=True, help="出力先 (.xlsx)")
    p.add_argument("--gantt-start", required=True, metavar="YYYY-MM-DD",
                   help="ガントチャート開始日")
    p.add_argument("--gantt-end",   required=True, metavar="YYYY-MM-DD",
                   help="ガントチャート終了日")

    return ap


def main():
    args = _build_parser().parse_args()
    {
        "add-time-cols":  cmd_add_time_cols,
        "add-status-col": cmd_add_status_col,
        "wbs2outlook":    cmd_wbs2outlook,
        "outlook2wbs":    cmd_outlook2wbs,
        "new-wbs":        cmd_new_wbs,
    }[args.command](args)


if __name__ == "__main__":
    main()
